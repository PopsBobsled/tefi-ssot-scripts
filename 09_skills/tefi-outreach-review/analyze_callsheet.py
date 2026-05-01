"""
TEFI Direct Outreach Call Sheet Analyzer — v2
Canonical column structure: A-J (research), K-P (engagement).
Usage: python3 analyze_callsheet.py <path_to_xlsx> [--registry <path>]
"""

import sys
import json
import argparse
import datetime
import pathlib
import pandas as pd

# ── Keyword lists for Notes column classification ─────────────────────────────
VISA_KEYWORDS        = ['visa', 'sponsorship', 'right to work', 'residency', 'citizen',
                        'overseas', 'accredited employer', 'work rights', 'working rights']
POSITIVE_KEYWORDS    = ['well-received', 'please apply', 'send in your email',
                        'attach the position description', 'kia ora', 'thank you for your interest',
                        'you could be considered', 'keen to see']
OVERQUALIFIED_KW     = ['overqualified', 'ovequalified', 'intermediate role',
                        'local applications', 'enough local', '200 applicants', 'locals preferred']
CLEARANCE_KW         = ['clearance']
REDIRECTED_KW        = ['apply online', 'apply through seek', 'apply via',
                        'recruitment portal', 'apply on our website', 'apply for the role through']
REJECTION_KW         = ['not progressing', 'regret', 'unable to proceed', 'not considering',
                        'unable to offer', 'cannot help', 'not registered', 'cannot support',
                        'not in a position']
INTERVIEW_KW         = ['interview', 'video call', 'face to face', 'meeting booked', 'zoom call']


def classify_notes(text):
    """Classify the free-text Notes / Outcome column."""
    if pd.isna(text) or str(text).strip() in ('', 'nan'):
        return 'no_notes'
    t = str(text).lower()
    if any(k in t for k in VISA_KEYWORDS):
        return 'visa_block'
    if any(k in t for k in OVERQUALIFIED_KW):
        return 'overqualified'
    if any(k in t for k in CLEARANCE_KW):
        return 'security_clearance'
    if any(k in t for k in INTERVIEW_KW):
        return 'interview_signal'
    if any(k in t for k in POSITIVE_KEYWORDS) and not any(k in t for k in REJECTION_KW):
        return 'positive'
    if any(k in t for k in REDIRECTED_KW):
        return 'redirected'
    if any(k in t for k in REJECTION_KW):
        return 'rejected'
    return 'other'


def normalise_sector(text):
    if pd.isna(text):
        return 'Unknown'
    t = str(text).lower()
    if any(k in t for k in ['charity', 'nfp', 'not for profit', 'not-for-profit', 'ngo', 'volunteer']):
        return 'NFP/Charity'
    if any(k in t for k in ['ministry', 'government', 'defence', 'defense', 'local authority', 'council']):
        return 'Government/Ministry'
    if any(k in t for k in ['health', 'te whatu', 'hospital', 'medical']):
        return 'Health'
    if 'university' in t or 'academic' in t:
        return 'University'
    if 'international' in t:
        return 'International Org'
    return 'Other/Unknown'


def normalise_region(text):
    if pd.isna(text) or str(text).strip() == '':
        return 'Unknown'
    t = str(text).lower().strip()
    if t.startswith('http') or t.startswith('www'):
        return 'Unknown'
    if 'auckland' in t or 'greenlane' in t or 'north shore' in t:
        return 'Auckland'
    if 'wellington' in t:
        return 'Wellington'
    if 'christchurch' in t:
        return 'Christchurch'
    if 'hamilton' in t:
        return 'Hamilton'
    if 'dunedin' in t:
        return 'Dunedin'
    if 'gisborne' in t:
        return 'Gisborne'
    if any(k in t for k in ['tasman', 'richmond', 'golden bay', 'nelson']):
        return 'Top of South Island'
    if any(k in t for k in ['wairarapa', 'masterton']):
        return 'Wairarapa'
    if any(k in t for k in ['anywhere', 'remote', 'flexible']):
        return 'Remote/Flexible'
    return 'Other'


def l_level(row):
    """
    Canonical TEFI L-levels based on new column structure.
    L1  — no call made (call_date empty)
    L2  — call attempted, not connected (Voicemail / Not Connected / Wrong Number)
    L3  — Live Connected, no response received yet
    L3b — Live Connected AND Response Received = Yes
    L4+ — Application Sent = Yes OR interview signal in notes
    """
    outcome  = str(row.get('call_outcome', '')).strip()
    response = str(row.get('response_received', '')).strip().lower()
    app_sent = str(row.get('application_sent', '')).strip().lower()
    notes    = str(row.get('notes', '')).lower()
    has_call = pd.notna(row.get('call_date')) and str(row.get('call_date', '')).strip() not in ('', 'nan')

    if any(k in notes for k in INTERVIEW_KW) or app_sent == 'yes':
        return 'L4+'
    if outcome == 'Live Connected':
        if response == 'yes':
            return 'L3b'
        return 'L3'
    if has_call and outcome in ('Voicemail', 'Not Connected', 'Wrong Number'):
        return 'L2'
    # Response received via email only (no live call recorded)
    if response == 'yes' and not has_call:
        return 'L2'
    return 'L1'


def weighted_avg(level_counts, total):
    if total == 0:
        return 0
    weights = {'L1': 1, 'L2': 2, 'L3': 3, 'L3b': 3.5, 'L4+': 4}
    score = sum(weights.get(k, 1) * v for k, v in level_counts.items())
    return round(score / total, 2)


# ── Weekly Tracker write ───────────────────────────────────────────────────────

def write_tracker_row(xlsx_path, metrics, today_str):
    """Append a snapshot row to the hidden 'Weekly Tracker' sheet."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        sheet_name = 'Weekly Tracker' if 'Weekly Tracker' in wb.sheetnames else None
        if sheet_name is None:
            return False
        ws = wb[sheet_name]
        # Find next empty row at/after row 2
        next_row = 2
        for r in range(2, (ws.max_row or 2) + 2):
            if ws.cell(row=r, column=1).value in (None, ''):
                next_row = r
                break
        row_values = [
            today_str,
            metrics.get('total_roles', 0),
            metrics.get('connected_calls', 0),
            metrics.get('responses', 0),
            metrics.get('applications_sent', 0),
            metrics.get('followups_sent', 0),
            metrics.get('weighted_avg_l', 0),
        ]
        for col_idx, val in enumerate(row_values, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)
        wb.save(xlsx_path)
        return True
    except Exception:
        return False


# ── Registry ──────────────────────────────────────────────────────────────────

def client_key_from_path(xlsx_path):
    return pathlib.Path(xlsx_path).stem.replace(' ', '_')


def load_registry(registry_path):
    try:
        with open(registry_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if 'clients' not in data:
                data['clients'] = {}
            return data
    except Exception:
        return {'clients': {}}


def save_registry(registry_path, data):
    try:
        p = pathlib.Path(registry_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        with open(p, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
        return True
    except Exception:
        return False


def snapshot_for_registry(metrics):
    keep = ['total_roles', 'contact_known', 'connected_calls', 'responses',
            'applications_sent', 'followups_sent', 'weighted_avg_l', 'analysis_date']
    return {k: metrics[k] for k in keep if k in metrics}


def compute_delta(current, previous):
    if previous is None:
        return None
    def d(key):
        return round((current.get(key) or 0) - (previous.get(key) or 0), 2)
    return {
        'delta_roles':        d('total_roles'),
        'delta_contacts':     d('contact_known'),
        'delta_connected':    d('connected_calls'),
        'delta_responses':    d('responses'),
        'delta_applications': d('applications_sent'),
        'delta_weighted_l':   d('weighted_avg_l'),
        'previous_date':      previous.get('analysis_date'),
    }


# ── Version detection ─────────────────────────────────────────────────────────

def detect_version(cs_sheet):
    """
    Detect call sheet version from raw sheet data.
    v2: 16 columns (A-P), 3 header rows, structured dropdowns.
    v1: 15 columns (A-O), 2 header rows, free-text comments.
    Returns ('v2', 3) or ('v1', 2) — (version_label, header_rows).
    """
    ncols = cs_sheet.shape[1]
    if ncols >= 16:
        return 'v2', 3
    return 'v1', 2


def build_v1_metrics(df):
    """
    Legacy column mapping for v1 call sheets.
    K=voice_call(Y/N), L=email_date, M=followup_date, N=comments, O=interested.
    Returns a partial metrics dict covering what v1 data supports.
    """
    PHONE_KW = ['voicemail', 'voice mail', 'not going through', 'went to voicemail',
                'left a message', 'left message', 'call not going']

    def classify_v1(text):
        if pd.isna(text) or str(text).strip() in ('', 'nan'):
            return 'no_response'
        t = str(text).lower()
        if any(k in t for k in VISA_KEYWORDS):        return 'visa_block'
        if any(k in t for k in PHONE_KW):             return 'phone_attempt'
        if any(k in t for k in OVERQUALIFIED_KW):     return 'overqualified'
        if any(k in t for k in CLEARANCE_KW):         return 'security_clearance'
        if any(k in t for k in INTERVIEW_KW):         return 'interview_signal'
        if any(k in t for k in POSITIVE_KEYWORDS) \
           and not any(k in t for k in REJECTION_KW): return 'positive'
        if any(k in t for k in REDIRECTED_KW):        return 'redirected'
        if any(k in t for k in REJECTION_KW):         return 'rejected'
        return 'other'

    def v1_l_level(row):
        comments = str(row.get('comments', '')).lower()
        voice    = str(row.get('voice_call', '')).strip().upper()
        rt       = row.get('response_type', '')
        if any(k in comments for k in INTERVIEW_KW):  return 'L4+'
        if voice == 'Y':                               return 'L3b'
        if rt == 'phone_attempt':                      return 'L3'
        if rt not in ('no_response', None, ''):        return 'L2'
        return 'L1'

    df['response_type'] = df['comments'].apply(classify_v1) if 'comments' in df.columns else 'no_response'
    df['notes_type']    = df['response_type']  # alias so downstream code works

    voice_confirmed = int((df.get('voice_call', pd.Series(dtype=str))
                           .astype(str).str.strip().str.upper() == 'Y').sum())
    email_dated     = int(df['email_date'].notna().sum()) if 'email_date' in df.columns else 0
    has_response    = int((df['response_type'] != 'no_response').sum())
    notes_counts    = df['response_type'].value_counts().to_dict()
    visa_block      = notes_counts.get('visa_block', 0)

    df['l_level'] = df.apply(v1_l_level, axis=1)

    return {
        'connected_calls':   voice_confirmed,   # best proxy in v1
        'voicemail_calls':   notes_counts.get('phone_attempt', 0),
        'not_connected':     0,
        'calls_attempted':   email_dated,
        'calls_last_5_days': None,              # not derivable from v1 data
        'responses':         has_response,
        'applications_sent': 0,                 # not tracked in v1
        'followups_sent':    0,                 # not tracked in v1
        'conversation_rate': None,
        'followthrough_rate':None,
        'notes_counts':      notes_counts,
        'visa_block_count':  visa_block,
        'visa_block_pct':    round(visa_block / has_response * 100) if has_response else 0,
        'l_level_col':       df['l_level'],
    }


# ── Core analysis ─────────────────────────────────────────────────────────────

def analyze(path, registry_path=None):
    today = datetime.date.today()
    today_iso   = today.isoformat()
    today_human = today.strftime('%d %b %Y')

    xl = pd.read_excel(path, sheet_name=None, header=None)
    sheet_names = list(xl.keys())

    # Prefer "Call Sheet" tab; fall back to first sheet
    cs_raw = xl.get('Call Sheet', xl[sheet_names[0]])

    # ── Version detection ─────────────────────────────────────────────────────
    sheet_version, header_rows = detect_version(cs_raw)

    df = cs_raw.iloc[header_rows:].copy()
    df.columns = range(len(df.columns))

    # Base research columns — same in both versions
    base_map = {
        0: 'job_link', 1: 'position', 2: 'company', 3: 'company_type',
        4: 'location', 5: 'contact_person', 6: 'contact_title',
        7: 'phone', 8: 'email', 9: 'website',
    }

    if sheet_version == 'v2':
        col_map = {**base_map,
            10: 'call_date', 11: 'call_outcome', 12: 'response_received',
            13: 'application_sent', 14: 'followup_date', 15: 'notes',
        }
    else:
        col_map = {**base_map,
            10: 'voice_call', 11: 'email_date', 12: 'followup_date',
            13: 'comments', 14: 'interested',
        }

    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
    df = df[df['position'].notna() | df['company'].notna()].reset_index(drop=True)

    total = len(df)

    # ── Core counts (research columns — same both versions) ───────────────────
    contact_known = int(df['contact_person'].notna().sum())
    phone_known   = int(df['phone'].notna().sum())
    email_known   = int(df['email'].notna().sum())

    # ── Engagement metrics — version-branched ─────────────────────────────────
    if sheet_version == 'v2':
        co = df.get('call_outcome', pd.Series(dtype=str)).astype(str).str.strip()
        connected_calls   = int((co == 'Live Connected').sum())
        voicemail_calls   = int((co == 'Voicemail').sum())
        not_connected     = int(co.isin(['Not Connected', 'Wrong Number']).sum())
        calls_attempted   = int(df['call_date'].notna().sum())

        mr = df.get('response_received', pd.Series(dtype=str)).astype(str).str.strip().str.lower()
        responses         = int((mr == 'yes').sum())

        na_ = df.get('application_sent', pd.Series(dtype=str)).astype(str).str.strip().str.lower()
        applications_sent = int((na_ == 'yes').sum())
        followups_sent    = int(df['followup_date'].notna().sum()) if 'followup_date' in df.columns else 0

        conversation_rate  = round(responses / connected_calls * 10, 2) if connected_calls else None
        followthrough_rate = round(applications_sent / responses * 10, 2) if responses else None

        df['notes_type'] = df['notes'].apply(classify_notes) if 'notes' in df.columns else 'no_notes'
        notes_counts     = df['notes_type'].value_counts().to_dict()
        visa_block_count = notes_counts.get('visa_block', 0)
        visa_block_pct   = round(visa_block_count / responses * 100) if responses else 0

        df['l_level']    = df.apply(l_level, axis=1)

        calls_last_5 = 0
        try:
            dates    = pd.to_datetime(df['call_date'], errors='coerce')
            cutoff   = pd.Timestamp(today) - pd.Timedelta(days=5)
            calls_last_5 = int(((co == 'Live Connected') & (dates >= cutoff)).sum())
        except Exception:
            pass

    else:
        # v1 — keyword-based estimation
        v1 = build_v1_metrics(df)
        connected_calls    = v1['connected_calls']
        voicemail_calls    = v1['voicemail_calls']
        not_connected      = v1['not_connected']
        calls_attempted    = v1['calls_attempted']
        calls_last_5       = v1['calls_last_5_days']
        responses          = v1['responses']
        applications_sent  = v1['applications_sent']
        followups_sent     = v1['followups_sent']
        conversation_rate  = v1['conversation_rate']
        followthrough_rate = v1['followthrough_rate']
        notes_counts       = v1['notes_counts']
        visa_block_count   = v1['visa_block_count']
        visa_block_pct     = v1['visa_block_pct']
        df['l_level']      = v1['l_level_col']

    # ── Lead highlights ───────────────────────────────────────────────────────
    notes_col = 'notes' if sheet_version == 'v2' else 'comments'

    def best_lead():
        pos = df[df['notes_type'].isin(['positive', 'interview_signal'])]
        if sheet_version == 'v2':
            pos = pos[pos.get('application_sent', pd.Series(dtype=str)).astype(str).str.lower() != 'yes']
        for _, r in pos.iterrows():
            return {'company': str(r.get('company', '')),
                    'position': str(r.get('position', '')),
                    'notes': str(r.get(notes_col, ''))[:120]}
        return None

    def untouched_lead():
        call_col = 'call_date' if sheet_version == 'v2' else 'email_date'
        if call_col not in df.columns:
            return None
        mask = (df['contact_person'].notna() & df['phone'].notna() &
                df[call_col].isna())
        for _, r in df[mask].iterrows():
            return {'company': str(r.get('company', '')),
                    'position': str(r.get('position', '')),
                    'contact': str(r.get('contact_person', ''))}
        return None

    def response_no_app():
        if sheet_version != 'v2':
            return None   # not reliably derivable from v1 data
        mask = (df['response_received'].astype(str).str.lower() == 'yes') & \
               (df['application_sent'].astype(str).str.lower() != 'yes')
        count = int(mask.sum())
        if count == 0:
            return None
        example = None
        for _, r in df[mask].iterrows():
            example = {'company': str(r.get('company', '')),
                       'position': str(r.get('position', ''))}
            break
        return {'count': count, 'example': example}

    # ── Weekly Tracker write ──────────────────────────────────────────────────
    metrics = {
        'analysis_date':       today_iso,
        'sheet_version':       sheet_version,
        'total_roles':         total,
        'contact_known':       contact_known,
        'contact_known_pct':   round(contact_known / total * 100) if total else 0,
        'phone_known':         phone_known,
        'email_known':         email_known,
        'calls_attempted':     calls_attempted,
        'connected_calls':     connected_calls,
        'voicemail_calls':     voicemail_calls,
        'not_connected':       not_connected,
        'calls_last_5_days':   calls_last_5,
        'responses':           responses,
        'applications_sent':   applications_sent,
        'followups_sent':      followups_sent,
        'conversation_rate':   conversation_rate,
        'followthrough_rate':  followthrough_rate,
        'visa_block_count':    visa_block_count,
        'visa_block_pct':      visa_block_pct,
        'notes_counts':        notes_counts,
        'sector_counts':       df['sector'].value_counts().to_dict() if 'sector' in df.columns else {},
        'region_counts':       df['region'].value_counts().to_dict() if 'region' in df.columns else {},
        'l_counts':            df['l_level'].value_counts().to_dict(),
        'weighted_avg_l':      weighted_avg(df['l_level'].value_counts().to_dict(), total),
        'best_lead':           best_lead(),
        'untouched_lead':      untouched_lead(),
        'response_no_app':     response_no_app(),
    }

    # Sector / region (ensure columns exist)
    if 'sector' not in df.columns:
        df['sector'] = df['company_type'].apply(normalise_sector)
    if 'region' not in df.columns:
        df['region'] = df['location'].apply(normalise_region)
    metrics['sector_counts'] = df['sector'].value_counts().to_dict()
    metrics['region_counts'] = df['region'].value_counts().to_dict()

    tracker_row_written = write_tracker_row(path, metrics, today_human)
    metrics['tracker_row_written'] = tracker_row_written

    # ── Registry ──────────────────────────────────────────────────────────────
    previous_snapshot = None
    delta = None
    registry_written = False
    client_key = client_key_from_path(path)

    if registry_path:
        reg = load_registry(registry_path)
        history = reg['clients'].get(client_key, [])
        if history:
            previous_snapshot = history[-1]
            delta = compute_delta(metrics, previous_snapshot)
        snap = snapshot_for_registry(metrics)
        snap['analysis_date'] = today_iso
        history.append(snap)
        reg['clients'][client_key] = history
        registry_written = save_registry(registry_path, reg)

    metrics['client_key']        = client_key
    metrics['registry_path']     = registry_path
    metrics['registry_written']  = registry_written
    metrics['previous_snapshot'] = previous_snapshot
    metrics['delta']             = delta

    return metrics


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx_path', help='Path to the TEFI call sheet .xlsx file')
    parser.add_argument('--registry', default=None, help='Path to tefi-registry.json')
    args = parser.parse_args()
    result = analyze(args.xlsx_path, registry_path=args.registry)
    print(json.dumps(result, indent=2, default=str))
